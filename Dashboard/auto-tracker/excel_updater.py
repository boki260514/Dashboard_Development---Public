"""
엑셀 트래킹 시트 업데이트 모듈.

DailyReport 객체를 받아서 엑셀 파일의 정확한 셀에 입력합니다.
업데이트 전 자동 백업, 덮어쓰기 경고를 포함합니다.
"""
from __future__ import annotations

import logging
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    BACKUP_FOLDER_NAME,
    DATA_START_COL,
    DAY_ROW,
    EXCEL_PATH,
    MONTH_ROW,
    ROW_MAPPING,
    ROW_OFFSET,
)
from parser import DailyReport, WorkItem


logger = logging.getLogger(__name__)


# 영문 월 이름 → 숫자
MONTH_NAME_TO_NUM = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12,
}


@dataclass
class UpdateResult:
    """단일 셀 업데이트 결과."""
    sheet: str
    row_label: str
    field_label: str  # actual / downtime / labor
    cell: str         # 예: "AG11"
    old_value: object
    new_value: object
    overwritten: bool = False  # 기존 값이 있어서 덮어쓴 경우


@dataclass
class UpdateReport:
    """전체 업데이트 결과 보고서."""
    sheet_name: str
    work_date: Optional[date]
    backup_path: Optional[str] = None
    results: list[UpdateResult] = None
    skipped: list[str] = None  # 매칭 실패한 항목들
    errors: list[str] = None

    def __post_init__(self) -> None:
        if self.results is None:
            self.results = []
        if self.skipped is None:
            self.skipped = []
        if self.errors is None:
            self.errors = []


def make_backup(excel_path: str) -> str:
    """
    엑셀 파일을 같은 폴더의 backups/ 안에 시각 포함 이름으로 복사합니다.

    Args:
        excel_path: 원본 엑셀 파일 경로.

    Returns:
        생성된 백업 파일의 절대 경로.
    """
    src = Path(excel_path)
    backup_dir = src.parent / BACKUP_FOLDER_NAME
    backup_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{src.stem}_backup_{timestamp}{src.suffix}"
    backup_path = backup_dir / backup_name
    shutil.copy2(src, backup_path)
    logger.info("백업 생성: %s", backup_path)
    return str(backup_path)


def find_date_column(
    ws: Worksheet,
    target_date: date,
) -> Optional[int]:
    """
    워크시트에서 특정 날짜에 해당하는 컬럼 번호를 찾습니다.

    7행에 월 이름(April / May / ...), 8행에 일자(1, 2, ...)가 있는 구조에서,
    target_date와 일치하는 셀의 컬럼을 찾아 반환합니다.

    Args:
        ws: openpyxl 워크시트.
        target_date: 찾을 날짜.

    Returns:
        해당 날짜의 컬럼 번호 (1부터). 못 찾으면 None.
    """
    target_month = target_date.month
    target_day = target_date.day

    # 현재 활성화된 월 컬럼 추적
    current_month = None
    for col in range(DATA_START_COL, ws.max_column + 1):
        # 7행에 월 이름이 있으면 갱신
        v_month = ws.cell(row=MONTH_ROW, column=col).value
        if v_month in MONTH_NAME_TO_NUM:
            current_month = MONTH_NAME_TO_NUM[v_month]

        # 8행의 일자 확인
        v_day = ws.cell(row=DAY_ROW, column=col).value
        if (current_month == target_month and
                isinstance(v_day, (int, float)) and
                int(v_day) == target_day):
            return col

    return None


def find_row_for_label(
    ws: Worksheet,
    row_label: str,
) -> Optional[int]:
    """
    워크시트의 B열에서 row_label과 일치하는 행 번호를 찾습니다.

    Args:
        ws: openpyxl 워크시트.
        row_label: 찾을 행 라벨 (예: "PB#01 Gap guarding Installation").

    Returns:
        해당 행 번호. 못 찾으면 None.
    """
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=2).value
        if v == row_label:
            return row
    return None


def map_item_to_row_label(
    sheet_name: str,
    item: WorkItem,
) -> Optional[str]:
    """
    카톡의 작업 항목을 엑셀의 행 라벨로 매핑합니다.

    config.py의 ROW_MAPPING에서 시트명과 정규화된 키로 조회합니다.

    Args:
        sheet_name: 시트명 (예: "CHW1FC").
        item: 작업 항목.

    Returns:
        엑셀 행 라벨. 못 찾으면 None.
    """
    sheet_mapping = ROW_MAPPING.get(sheet_name, {})
    # 라인 + 작업종류 조합 키 시도
    key = f"{item.line_normalized}_{item.work_type}"
    if key in sheet_mapping:
        return sheet_mapping[key]
    return None


def update_excel(
    report: DailyReport,
    excel_path: str = EXCEL_PATH,
    create_backup: bool = True,
    overwrite: bool = False,
) -> UpdateReport:
    """
    DailyReport 내용을 엑셀 파일에 입력합니다.

    Args:
        report: 입력할 일일 보고서.
        excel_path: 엑셀 파일 경로.
        create_backup: True면 입력 전 백업 생성.
        overwrite: True면 기존 값을 무조건 덮어씀.
                   False면 기존 값이 있을 때 경고하고 덮어씀(로그에 표시).

    Returns:
        UpdateReport 객체. 입력 결과와 누락/에러를 담음.
    """
    update = UpdateReport(
        sheet_name=report.site_sheet,
        work_date=report.work_date,
    )

    # 백업
    if create_backup:
        try:
            update.backup_path = make_backup(excel_path)
        except Exception as exc:
            update.errors.append(f"백업 실패: {exc}")
            return update

    # 워크북 로드
    try:
        wb = load_workbook(excel_path)
    except Exception as exc:
        update.errors.append(f"엑셀 파일 로드 실패: {exc}")
        return update

    # 시트 선택
    if report.site_sheet not in wb.sheetnames:
        update.errors.append(
            f"시트 '{report.site_sheet}'를 찾을 수 없습니다. "
            f"(존재하는 시트: {wb.sheetnames})"
        )
        return update
    ws = wb[report.site_sheet]

    # 날짜 → 컬럼 매핑
    if report.work_date is None:
        update.errors.append(
            "메시지에서 작업 일자를 찾을 수 없습니다."
        )
        return update
    date_col = find_date_column(ws, report.work_date)
    if date_col is None:
        update.errors.append(
            f"엑셀 시트에 {report.work_date} 날짜 컬럼이 없습니다."
        )
        return update
    date_col_letter = get_column_letter(date_col)

    # 각 작업 항목 입력
    for item in report.items:
        row_label = map_item_to_row_label(report.site_sheet, item)
        if not row_label:
            update.skipped.append(
                f"매핑 실패: '{item.line_raw}' + '{item.work_type_raw}' "
                f"→ ROW_MAPPING에 등록 필요"
            )
            continue

        row_num = find_row_for_label(ws, row_label)
        if row_num is None:
            update.skipped.append(
                f"행 헤더 '{row_label}'를 시트 {report.site_sheet}에서 "
                "찾을 수 없습니다."
            )
            continue

        # Actual 행에 수량 입력
        actual_row = row_num + ROW_OFFSET["actual"]
        cell = ws.cell(row=actual_row, column=date_col)
        old_value = cell.value
        was_existing = old_value not in (None, "", 0)
        cell.value = item.quantity
        update.results.append(UpdateResult(
            sheet=report.site_sheet,
            row_label=row_label,
            field_label="actual",
            cell=f"{date_col_letter}{actual_row}",
            old_value=old_value,
            new_value=item.quantity,
            overwritten=was_existing,
        ))

        # Labor 행에 분배된 인원 입력 (item.workers 우선, 없으면 총원)
        labor_value = None
        if item.workers > 0:
            labor_value = item.workers
        elif report.total_workers is not None:
            labor_value = report.total_workers
        if labor_value is not None:
            labor_row = row_num + ROW_OFFSET["labor"]
            cell = ws.cell(row=labor_row, column=date_col)
            old_value = cell.value
            was_existing = old_value not in (None, "", 0)
            cell.value = labor_value
            update.results.append(UpdateResult(
                sheet=report.site_sheet,
                row_label=row_label,
                field_label="labor",
                cell=f"{date_col_letter}{labor_row}",
                old_value=old_value,
                new_value=labor_value,
                overwritten=was_existing,
            ))

        # Downtime 행 위치 (수식에서 참조하기 위해 항상 계산)
        downtime_row = row_num + ROW_OFFSET["downtime"]

        # Downtime 행에 작업시간 입력
        if report.work_hours is not None:
            cell = ws.cell(row=downtime_row, column=date_col)
            old_value = cell.value
            was_existing = old_value not in (None, "", 0)
            cell.value = report.work_hours
            update.results.append(UpdateResult(
                sheet=report.site_sheet,
                row_label=row_label,
                field_label="downtime",
                cell=f"{date_col_letter}{downtime_row}",
                old_value=old_value,
                new_value=report.work_hours,
                overwritten=was_existing,
            ))

        # Installation/Hour 행에 수식 입력 (Actual ÷ Downtime)
        # Downtime이 0일 때 #DIV/0! 방지를 위해 IFERROR 또는 IF 사용
        installation_hour_row = row_num + ROW_OFFSET["installation_hour"]
        actual_cell_ref = f"{date_col_letter}{actual_row}"
        downtime_cell_ref = f"{date_col_letter}{downtime_row}"
        formula = (
            f'=IFERROR({actual_cell_ref}/{downtime_cell_ref},0)'
        )
        cell = ws.cell(row=installation_hour_row, column=date_col)
        old_value = cell.value
        was_existing = old_value not in (None, "", 0)
        cell.value = formula
        update.results.append(UpdateResult(
            sheet=report.site_sheet,
            row_label=row_label,
            field_label="install/hr",
            cell=f"{date_col_letter}{installation_hour_row}",
            old_value=old_value,
            new_value=formula,
            overwritten=was_existing,
        ))

    # 저장
    try:
        wb.save(excel_path)
    except PermissionError:
        update.errors.append(
            "엑셀 파일에 쓸 수 없습니다. "
            "파일이 다른 프로그램에서 열려있는지 확인하세요."
        )
        return update
    except Exception as exc:
        update.errors.append(f"저장 실패: {exc}")
        return update

    return update

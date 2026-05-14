"""
Mail_Log.xlsx 스키마와 빈 템플릿 생성기.

상수만 모아둔 모듈. 다른 모듈에서 import 해서 사용한다.

수동 실행:
    python mail_log_schema.py            # Mail_Log.xlsx 파일이 없으면 생성
    python mail_log_schema.py --force    # 있어도 덮어쓰기 (주의: 데이터 손실)
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

# Mail_Log.xlsx 위치: auto-tracker의 부모 폴더 (= Dashboard_Development - Dashboard)
_HERE = Path(__file__).parent
MAIL_LOG_XLSX = _HERE.parent / "Mail_Log.xlsx"

# 메일 로그 대시보드 HTML 출력
MAIL_DASHBOARD_HTML = _HERE / "mail_dashboard.html"

# 상태 파일 (마지막 성공 실행 시각 등)
MAIL_STATE_JSON = _HERE / "mail_state.json"


# ============================================================
# Mail_Log 시트 컬럼 정의 (순서가 곧 컬럼 위치)
# ============================================================
MAIL_LOG_SHEET = "Mail_Log"

MAIL_LOG_COLUMNS = [
    "메일ID",
    "수신일시",
    "발신자이름",
    "발신자이메일",
    "회사명",
    "제목",
    "업무유형",
    "핵심내용요약",
    "요청사항",
    "금액",
    "수량",
    "납기일",
    "담당자",
    "우선순위",
    "상태",
    "첨부파일명",
    "원문참조",
    "처리일시",
    "신뢰도",
    "비고",
]

# 컬럼명 → 인덱스(0-based) 빠른 조회용
MAIL_LOG_COL_INDEX = {name: i for i, name in enumerate(MAIL_LOG_COLUMNS)}

# 허용값(드롭다운 후보)
WORK_TYPES = ["견적", "주문", "계약", "정산", "CS", "일정", "보고", "기타"]
PRIORITIES = ["긴급", "높음", "보통", "낮음"]
STATUSES = ["신규", "진행중", "검토필요", "완료"]
CONFIDENCES = ["높음", "중간", "낮음"]


# ============================================================
# Run_Log 시트 컬럼 정의
# ============================================================
RUN_LOG_SHEET = "Run_Log"

RUN_LOG_COLUMNS = [
    "실행ID",
    "시작시각",
    "종료시각",
    "확인메일수",
    "신규반영",
    "중복제외",
    "검토필요",
    "비업무제외",
    "오류여부",
    "오류상세",
    "대시보드상태",
]


# ============================================================
# 빈 템플릿 생성
# ============================================================
def create_empty_template(target_path: Path | str = MAIL_LOG_XLSX,
                         force: bool = False) -> Path:
    """Mail_Log.xlsx를 빈 템플릿으로 생성한다. 이미 있으면 force=True일 때만 덮어쓴다."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    target_path = Path(target_path)
    if target_path.exists() and not force:
        print(f"[SKIP] 이미 존재: {target_path}")
        return target_path

    wb = Workbook()

    # 기본 시트 → Mail_Log로 변경
    ws = wb.active
    ws.title = MAIL_LOG_SHEET

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")

    # Mail_Log 헤더
    for col_idx, name in enumerate(MAIL_LOG_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 너비 (대충 합리적인 기본값)
    widths = {
        "메일ID": 22, "수신일시": 19, "발신자이름": 14, "발신자이메일": 26,
        "회사명": 16, "제목": 36, "업무유형": 10, "핵심내용요약": 40,
        "요청사항": 30, "금액": 13, "수량": 8, "납기일": 12,
        "담당자": 12, "우선순위": 9, "상태": 10, "첨부파일명": 26,
        "원문참조": 14, "처리일시": 19, "신뢰도": 8, "비고": 24,
    }
    for col_idx, name in enumerate(MAIL_LOG_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 14)

    ws.freeze_panes = "A2"

    # Run_Log 시트
    ws2 = wb.create_sheet(title=RUN_LOG_SHEET)
    for col_idx, name in enumerate(RUN_LOG_COLUMNS, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    run_widths = {
        "실행ID": 12, "시작시각": 19, "종료시각": 19, "확인메일수": 10,
        "신규반영": 10, "중복제외": 10, "검토필요": 10, "비업무제외": 11,
        "오류여부": 9, "오류상세": 40, "대시보드상태": 14,
    }
    for col_idx, name in enumerate(RUN_LOG_COLUMNS, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = run_widths.get(name, 14)
    ws2.freeze_panes = "A2"

    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    print(f"[OK] 생성됨: {target_path}")
    return target_path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--force", action="store_true",
                    help="이미 있어도 덮어쓰기 (데이터 손실 주의)")
    ap.add_argument("--path", default=str(MAIL_LOG_XLSX),
                    help="생성할 파일 경로")
    args = ap.parse_args()
    create_empty_template(args.path, force=args.force)
    return 0


if __name__ == "__main__":
    sys.exit(main())

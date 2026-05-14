"""
트래킹 시트(엑셀)에서 대시보드용 데이터를 추출하는 모듈.

엑셀 구조:
  - 시트명 = 현장 코드 (CHW1FC, ECH2FC 등)
  - B열 = 작업 항목명 ("PB#01 Gap guarding Installation" 등)
  - 7행 = 월 헤더 (April, May, ...)
  - 8행 = 일자 (1, 2, ..., 31)
  - 항목별 5행 그룹:
      Plan / Actual / Downtime window hours / Installation per Hour / Labor
  - CQ열 = 누적 합계 (Total)
  - CR열 = 진행률 (%)
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# 엑셀 위치 상수 (트래킹 시트와 일치)
DATA_START_COL = 4   # D열부터 데이터 시작
MONTH_ROW = 7
DAY_ROW = 8
TOTAL_COL_DEFAULT = 95   # CQ열 = 누적 총합 (4월 시작 시트)
# 주의: DON1FC/MCN1FC처럼 5월 시작 시트는 트래킹 기간이 92일이라
# Total열이 CR(96)로 한 칸 밀려있음. _find_total_col로 시트별 동적 탐지.
CARRIER_COL_NAME = "Total Completed Carriers"  # CT열 라벨 (시트마다 위치 다를 수 있음)


def _find_total_col(ws) -> int:
    """시트의 7행에서 'Total' 라벨을 가진 컬럼 번호 반환. 못찾으면 기본값(95)."""
    for c in range(80, min(ws.max_column + 1, 120)):
        v = ws.cell(MONTH_ROW, c).value
        if isinstance(v, str) and v.strip() == "Total":
            return c
    return TOTAL_COL_DEFAULT

MONTH_NAME_TO_NUM = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12,
}


@dataclass
class WorkItem:
    """단일 작업 항목 (예: PB#01 Gap guarding Installation)."""
    label: str              # "PB#01 Gap guarding Installation"
    short_name: str         # "PB#01"
    work_type: str          # "Gap guarding" / "Cable" / "Cover" / "Gap Plate"
    line_position: str = "" # "상단부" / "하단부" / "" (해당 없으면)
    plan_total: float = 0.0      # 전체 계획량
    actual_total: float = 0.0    # 누적 실적
    progress_pct: float = 0.0    # 진행률 (0.0~1.0)
    daily_data: list[tuple[date, float]] = field(default_factory=list)
    # 멘데이(Man-day) 관련
    used_mday_total: float = 0.0    # Labor 행 누적 사용 멘데이
    today_mday: float = 0.0         # 오늘 사용 멘데이
    # 최근 작업 정보
    last_date: Optional[date] = None
    last_qty: float = 0.0
    last_workers: int = 0
    last_install_hour: float = 0.0


@dataclass
class SiteData:
    """한 현장(시트)의 모든 데이터."""
    site_code: str          # "CHW1FC"
    items: list[WorkItem] = field(default_factory=list)

    # 집계 정보
    total_plan: float = 0.0
    total_actual: float = 0.0
    overall_pct: float = 0.0
    total_workers_used: int = 0   # 누적 인원 (모든 항목 Labor 합)
    total_workers_planned: int = 0  # 계획 인원 (모든 항목 Labor Plan 합)
    avg_daily_output: float = 0.0
    days_elapsed: int = 0
    days_total: int = 0
    days_remaining: int = 0

    # 멘데이(Man-day) 집계
    total_plan_mday: float = 0.0     # 총 계획 멘데이 (config.PLAN_MDAY_BY_SITE 값)
    total_used_mday: float = 0.0     # 현재까지 누적 사용 멘데이
    total_today_mday: float = 0.0    # 금일 작업시 사용한 멘데이

    # 이 현장의 일별 추이 (작업이 있던 날만, 시간순)
    daily_trend: list = field(default_factory=list)
    # 이 현장의 최근 활동 로그 (날짜 역순)
    recent_activity: list = field(default_factory=list)


@dataclass
class DashboardData:
    """대시보드 전체 데이터."""
    generated_at: datetime = field(default_factory=datetime.now)
    excel_path: str = ""
    sites: list[SiteData] = field(default_factory=list)
    # 활동 로그 (최근 n건, 모든 시트 통합)
    recent_activity: list[dict] = field(default_factory=list)
    # 일별 추이 (최근 14일, 모든 시트 통합)
    daily_trend: list[tuple[date, float]] = field(default_factory=list)


def _extract_short_name(label: str) -> tuple[str, str, str]:
    """
    'PB#01 Gap guarding Installation' →
        short='PB#01', work_type='Gap guarding', position=''
    
    'PB#02 Cover Installation' →
        short='PB#02', work_type='Cover', position=''
    """
    # work type 키워드들
    if "Gap guarding" in label:
        work_type = "Gap guarding"
    elif "Cable" in label:
        work_type = "Cable"
    elif "Gap Plate" in label:
        work_type = "Gap Plate"
    elif "Cover" in label:
        work_type = "Cover"
    else:
        work_type = "Installation"
    
    # short name = 첫 단어 (PB#01, ST#01, CB 등)
    parts = label.split()
    short = parts[0] if parts else label
    
    # 위치 (상단부/하단부) - 라벨에서 안 나오니까 별도 매핑 필요할 수 있음
    # 일단 빈 값으로
    position = ""
    
    return short, work_type, position


def _find_item_rows(ws) -> list[tuple[int, str]]:
    """
    시트에서 작업 항목들의 행 번호와 라벨을 찾아 반환.
    
    Returns:
        [(row_num, label), ...] 형식
    """
    items = []
    for row in range(8, ws.max_row + 1):
        b_val = ws.cell(row=row, column=2).value
        if b_val and isinstance(b_val, str) and "Installation" in b_val:
            # "Installation / Hour" 헤더는 제외
            if "/ Hour" in b_val or "Hour" in b_val.split()[-1]:
                continue
            items.append((row, b_val.strip()))
    return items


def _read_total_cell(ws, ws_formula, row: int, total_col: int = None) -> float:
    """
    Plan 행의 Total 열 (총 계획량)을 읽음.

    엑셀에 두 가지 패턴 가능:
      1) =395 같은 단순 숫자 → 캐시값 또는 수식에서 추출
      2) =SUM(D10:CP10) 같은 수식 → D~CP 직접 합산

    수식 캐시가 만료되어 있어도 정확한 값 반환.

    Args:
        ws: data_only=True 시트 (캐시값).
        ws_formula: data_only=False 시트 (수식 텍스트). None 가능.
        row: 읽을 행 번호.
        total_col: Total 열 번호. None이면 시트에서 동적 탐지.
                   (DON1FC/MCN1FC는 CR(96), 그 외는 CQ(95))
    """
    if total_col is None:
        total_col = _find_total_col(ws)

    # 1단계: 캐시된 값 시도
    cached = ws.cell(row=row, column=total_col).value
    if isinstance(cached, (int, float)) and cached != 0:
        return float(cached)

    # 2단계: 수식 시트에서 읽기 (있으면)
    if ws_formula is not None:
        formula = ws_formula.cell(row=row, column=total_col).value
        if isinstance(formula, (int, float)):
            return float(formula)
        if isinstance(formula, str) and formula.startswith("="):
            # 패턴 1: "=395" 같은 단순 숫자 → 추출
            formula_body = formula[1:].strip()
            try:
                return float(formula_body)
            except ValueError:
                pass
            # 패턴 2: SUM 수식 → 일별 셀 직접 합산
            # (이때는 캐시값 시트의 셀들을 더해야 정확함)

    # 3단계: 마지막 수단 — 일별 셀들 직접 합산 (D ~ Total열 직전까지)
    total = 0.0
    for col in range(DATA_START_COL, total_col):
        v = ws.cell(row=row, column=col).value
        if isinstance(v, (int, float)):
            total += float(v)
    return total


def _column_to_date(ws, col: int, year: int) -> Optional[date]:
    """
    워크시트의 컬럼 번호를 날짜로 변환.
    7행에서 가장 가까운 월 헤더 찾고, 8행에서 일자 읽기.
    """
    # 8행에서 일자
    day = ws.cell(row=DAY_ROW, column=col).value
    if not isinstance(day, (int, float)):
        return None
    
    # 7행을 거슬러 올라가며 가장 최근 월 헤더 찾기
    month = None
    for c in range(col, 0, -1):
        m = ws.cell(row=MONTH_ROW, column=c).value
        if m in MONTH_NAME_TO_NUM:
            month = MONTH_NAME_TO_NUM[m]
            break
    
    if month is None:
        return None
    
    try:
        return date(year, month, int(day))
    except ValueError:
        return None


def extract_site_data(ws, ws_formula=None, year: int = 2026) -> SiteData:
    """
    한 워크시트에서 SiteData 추출.
    
    Args:
        ws: openpyxl 워크시트 (data_only=True, 캐시값 읽기).
        ws_formula: openpyxl 워크시트 (data_only=False, 수식 텍스트 읽기).
                    None이면 ws만 사용.
        year: 데이터 연도.
    
    Returns:
        SiteData 객체.
    """
    site = SiteData(site_code=ws.title)

    # Total 열 동적 탐지 (시트마다 다를 수 있음: CQ 또는 CR)
    total_col = _find_total_col(ws)

    # 작업 항목 행들 찾기 (캐시값 시트에서)
    item_rows = _find_item_rows(ws)
    
    # 일자 컬럼 매핑
    date_col_map: dict[int, date] = {}
    for col in range(DATA_START_COL, ws.max_column + 1):
        d = _column_to_date(ws, col, year)
        if d:
            date_col_map[col] = d
    
    today = date.today()
    
    for row, label in item_rows:
        actual_row = row + 1  # Actual 행
        labor_row = row + 4   # Labor 행
        install_hour_row = row + 3  # Installation / Hour 행

        short, work_type, position = _extract_short_name(label)
        item = WorkItem(
            label=label,
            short_name=short,
            work_type=work_type,
            line_position=position,
        )

        # 일별 데이터 수집 (모든 데이터 컬럼)
        # 진행률 계산을 위해 0인 값도 포함해서 모든 일별 입력값을 모음
        all_daily_values = []
        for col, d in date_col_map.items():
            v = ws.cell(row=actual_row, column=col).value
            if isinstance(v, (int, float)):
                all_daily_values.append((d, float(v)))
                if v > 0:
                    item.daily_data.append((d, float(v)))

        # ====================================================
        # 멘데이(M-day) 계산 - Labor 행
        # ====================================================
        # Labor 행은 일별 투입 인원수가 저장됨.
        # 누적 사용 멘데이 = 모든 일자의 Labor 값 합산
        # 금일 사용 멘데이 = today에 해당하는 컬럼의 Labor 값
        used_mday_total = 0.0
        today_mday = 0.0
        for col, d in date_col_map.items():
            lv = ws.cell(row=labor_row, column=col).value
            if isinstance(lv, (int, float)) and lv > 0:
                used_mday_total += float(lv)
                if d == today:
                    today_mday += float(lv)
        item.used_mday_total = used_mday_total
        item.today_mday = today_mday
        
        # ====================================================
        # 누적 합계 직접 계산 (CQ열 수식 캐시 의존 X)
        # ====================================================
        # 엑셀의 CQ열은 =SUM(D11:CP11) 같은 수식이지만,
        # openpyxl이 수식을 직접 계산 안 하므로,
        # 우리가 직접 일별 값들을 모두 더해서 누적 합계 계산.
        actual_total_calc = sum(v for _, v in all_daily_values)
        
        # ====================================================
        # 계획 총량 계산 (CQ열, Plan 행)
        # ====================================================
        # Plan 행의 CQ열은 두 가지 패턴 가능:
        #   1) =395 같은 단순 숫자 → 그대로 사용
        #   2) =SUM(D10:CP10) 같은 수식 → 직접 계산
        # 캐시가 만료되어 있어도 안전하게 처리.
        plan_total = _read_total_cell(ws, ws_formula, row, total_col=total_col)
        
        item.plan_total = plan_total
        item.actual_total = actual_total_calc
        # 진행률도 직접 계산 (CR열 수식 의존 X)
        item.progress_pct = (
            actual_total_calc / plan_total if plan_total > 0 else 0.0
        )
        
        # 가장 최근 작업
        if item.daily_data:
            item.daily_data.sort(key=lambda x: x[0])
            last_d, last_q = item.daily_data[-1]
            item.last_date = last_d
            item.last_qty = last_q
            
            # 마지막 작업 날짜의 인원 + Install/Hour
            for col, d in date_col_map.items():
                if d == last_d:
                    labor = ws.cell(row=labor_row, column=col).value
                    install = ws.cell(row=install_hour_row, column=col).value
                    if isinstance(labor, (int, float)):
                        item.last_workers = int(labor)
                    if isinstance(install, (int, float)):
                        item.last_install_hour = float(install)
                    break
        
        site.items.append(item)
    
    # 시트 전체 집계
    site.total_plan = sum(it.plan_total for it in site.items)
    site.total_actual = sum(it.actual_total for it in site.items)
    site.overall_pct = (
        site.total_actual / site.total_plan if site.total_plan > 0 else 0.0
    )

    # 멘데이(M-day) 집계
    site.total_used_mday = sum(it.used_mday_total for it in site.items)
    site.total_today_mday = sum(it.today_mday for it in site.items)
    # total_plan_mday는 config에서 읽어 extract_dashboard_data에서 주입
    
    # 일별 추이 (모든 항목 통합)
    daily_combined: dict[date, float] = {}
    for item in site.items:
        for d, q in item.daily_data:
            daily_combined[d] = daily_combined.get(d, 0) + q
    
    sorted_dates = sorted(daily_combined.keys())
    if sorted_dates:
        # 가장 빠른 작업일 ~ 오늘 사이의 작업일 수
        first_work_day = sorted_dates[0]
        work_days = [d for d in sorted_dates if d <= today]
        if work_days:
            site.days_elapsed = len(work_days)
            total_qty = sum(daily_combined[d] for d in work_days)
            site.avg_daily_output = total_qty / max(len(work_days), 1)
    
    return site


def collect_recent_activity(sites: list[SiteData], limit: int = 30) -> list[dict]:
    """
    모든 시트에서 최근 작업 내역을 모아 시간 역순으로 정렬.
    
    Returns:
        [{date, site, item, qty, workers, install_hour}, ...]
    """
    all_records = []
    for site in sites:
        for item in site.items:
            for d, q in item.daily_data:
                all_records.append({
                    "date": d,
                    "site": site.site_code,
                    "item_label": item.label,
                    "short_name": item.short_name,
                    "work_type": item.work_type,
                    "qty": q,
                    "workers": 0,        # 일별 인원은 별도 추출 필요
                    "install_hour": 0,   # 일별 install/h
                })
    # 날짜 역순 정렬
    all_records.sort(key=lambda r: r["date"], reverse=True)
    return all_records[:limit]


def collect_daily_trend(sites: list[SiteData], days: Optional[int] = None) -> list[tuple[date, float]]:
    """
    일별 합계 추이.

    Args:
        days: None이면 작업이 있던 모든 날짜(전체 기간) 반환.
              정수면 최근 N일 (작업 없는 날도 0으로 포함).
    """
    daily_combined: dict[date, float] = {}
    for site in sites:
        for item in site.items:
            for d, q in item.daily_data:
                daily_combined[d] = daily_combined.get(d, 0) + q

    if not daily_combined:
        return []

    sorted_dates = sorted(daily_combined.keys())

    # days=None → 작업이 있던 모든 날짜를 시간순으로 반환 (전체 기간)
    if days is None:
        return [(d, daily_combined[d]) for d in sorted_dates]

    # days=정수 → 최근 N일 (오늘 기준, 빈 날 포함)
    today = date.today()
    end_date = max(sorted_dates[-1], today)
    from datetime import timedelta
    result = []
    for i in range(days - 1, -1, -1):
        d = end_date - timedelta(days=i)
        result.append((d, daily_combined.get(d, 0)))
    return result


def extract_dashboard_data(excel_path: str) -> DashboardData:
    """
    트래킹 시트 엑셀 파일에서 대시보드 데이터 전체 추출.
    
    Args:
        excel_path: 트래킹 엑셀 파일 경로.
    
    Returns:
        DashboardData 객체.
    """
    # 파일명에서 연도 추출 (예: "...2026.05.04.xlsx")
    year = 2026
    fname = Path(excel_path).name
    import re
    m = re.search(r"(\d{4})[._]", fname)
    if m:
        year = int(m.group(1))
    
    # 두 번 열어서 합치기:
    # - data_only=True: 캐시된 수식 결과값
    # - data_only=False: 실제 셀 값/수식 텍스트
    # Plan 행은 수식이 단순 숫자이거나 SUM 일 수 있어서 둘 다 봐야 함.
    wb_value = load_workbook(excel_path, data_only=True)
    wb_formula = load_workbook(excel_path, data_only=False)
    
    data = DashboardData(
        excel_path=excel_path,
        generated_at=datetime.now(),
    )
    
    # 현장별 총 계획 멘데이 매핑 (config에 정의)
    try:
        from config import PLAN_MDAY_BY_SITE
    except ImportError:
        PLAN_MDAY_BY_SITE = {}

    for sheet_name in wb_value.sheetnames:
        ws_value = wb_value[sheet_name]
        ws_formula = wb_formula[sheet_name]
        site = extract_site_data(ws_value, ws_formula, year=year)
        # 총 계획 멘데이 주입
        site.total_plan_mday = float(PLAN_MDAY_BY_SITE.get(sheet_name, 0) or 0)
        # 현장별 일별 추이 / 최근 활동 (사이트 1개만 넘겨서 계산)
        site.daily_trend = collect_daily_trend([site])
        site.recent_activity = collect_recent_activity([site], limit=30)
        data.sites.append(site)

    # 전체 합산(메인 대시보드 폴백용)
    data.recent_activity = collect_recent_activity(data.sites)
    data.daily_trend = collect_daily_trend(data.sites)

    return data


if __name__ == "__main__":
    pass

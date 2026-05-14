"""
이메일로 받은 'M-day Tracking Sheet' 첨부를 기존 HCNSSERVICE_Tracking_v2.xlsx에
안전하게 머지하는 도구.

머지 규칙 (사용자 스펙 "사람 입력값 덮어쓰기 금지" 준수):
    - 둘 다 빈 셀 → 무시
    - 첨부=값, 기존=빈 → 추가 (덮어쓰기 아님)
    - 첨부=값, 기존=같은 값 → 무시
    - 첨부=A, 기존=B (다름) → 충돌 리스트에만 기록, 자동 변경 없음
    - 첨부=빈, 기존=값 → 우리 데이터 보존

사용:
    python merge_tracker_from_email.py <첨부.xlsx>             # 비교만
    python merge_tracker_from_email.py <첨부.xlsx> --apply     # 안전 머지 적용 (백업 후)
"""
from __future__ import annotations

import argparse
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import EXCEL_PATH, BACKUP_FOLDER_NAME


@dataclass
class CellDiff:
    sheet: str
    cell: str            # 예: "AG12"
    old: object          # 기존 트래커 값
    new: object          # 첨부 값
    kind: str            # "add", "same", "conflict", "ours_only"


@dataclass
class DiffSummary:
    added: list[CellDiff] = field(default_factory=list)
    same: int = 0
    conflicts: list[CellDiff] = field(default_factory=list)
    ours_only: list[CellDiff] = field(default_factory=list)
    sheets_only_in_email: list[str] = field(default_factory=list)
    sheets_only_in_tracker: list[str] = field(default_factory=list)


def _norm(v):
    """비교용 정규화. None/빈문자열/공백은 모두 'empty'로 본다. 숫자는 float화."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    if isinstance(v, (int, float)):
        # 0과 비어있음은 구분! (0은 명시적 0)
        return float(v)
    return v


def _open_read_only(path: Path):
    """엑셀로 열려서 잠긴 경우 임시 사본을 만들어 그걸 읽는다."""
    try:
        return load_workbook(path, data_only=True, read_only=True), None
    except PermissionError:
        import tempfile, shutil as _sh
        tmp = Path(tempfile.gettempdir()) / f"_locked_{path.name}"
        _sh.copy2(path, tmp)
        print(f"  [잠김] {path.name} 사용 중 → 임시 사본으로 읽기: {tmp}")
        return load_workbook(tmp, data_only=True, read_only=True), tmp


def compare(email_path: Path, tracker_path: Path) -> DiffSummary:
    summary = DiffSummary()

    wb_e, _ = _open_read_only(email_path)
    wb_t, _ = _open_read_only(tracker_path)

    sheets_e = set(wb_e.sheetnames)
    sheets_t = set(wb_t.sheetnames)
    summary.sheets_only_in_email = sorted(sheets_e - sheets_t)
    summary.sheets_only_in_tracker = sorted(sheets_t - sheets_e)
    common = sorted(sheets_e & sheets_t)

    for sn in common:
        ws_e = wb_e[sn]
        ws_t = wb_t[sn]
        max_r = max(ws_e.max_row or 0, ws_t.max_row or 0)
        max_c = max(ws_e.max_column or 0, ws_t.max_column or 0)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                v_e = _norm(ws_e.cell(row=r, column=c).value)
                v_t = _norm(ws_t.cell(row=r, column=c).value)
                if v_e is None and v_t is None:
                    continue
                cell = f"{get_column_letter(c)}{r}"
                if v_e is not None and v_t is None:
                    summary.added.append(CellDiff(sn, cell, None,
                        ws_e.cell(row=r, column=c).value, "add"))
                elif v_e == v_t:
                    summary.same += 1
                elif v_e is None and v_t is not None:
                    summary.ours_only.append(CellDiff(sn, cell,
                        ws_t.cell(row=r, column=c).value, None, "ours_only"))
                else:
                    summary.conflicts.append(CellDiff(sn, cell,
                        ws_t.cell(row=r, column=c).value,
                        ws_e.cell(row=r, column=c).value, "conflict"))
    wb_e.close()
    wb_t.close()
    return summary


def apply_merge(email_path: Path, tracker_path: Path,
                summary: DiffSummary) -> Path:
    """summary.added 만 트래커에 추가 (덮어쓰기 없음). 백업 자동 생성."""
    backup_dir = tracker_path.parent / BACKUP_FOLDER_NAME
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{tracker_path.stem}_merge_{ts}{tracker_path.suffix}"
    shutil.copy2(tracker_path, backup_path)

    wb_t = load_workbook(tracker_path)
    for diff in summary.added:
        ws = wb_t[diff.sheet]
        col_letter = "".join(ch for ch in diff.cell if ch.isalpha())
        row_str = "".join(ch for ch in diff.cell if ch.isdigit())
        ws[diff.cell] = diff.new
    wb_t.save(tracker_path)
    wb_t.close()
    return backup_path


def print_summary(summary: DiffSummary, max_show: int = 20) -> None:
    print()
    print("=" * 70)
    print("머지 시뮬레이션 결과")
    print("=" * 70)
    print(f"  공통 시트 동일 셀 수    : {summary.same:,}")
    print(f"  자동 추가 후보(첨부→트래커 빈셀): {len(summary.added):,}")
    print(f"  충돌(값 다름, 자동 변경 X): {len(summary.conflicts):,}")
    print(f"  우리만 갖고 있는 셀     : {len(summary.ours_only):,}")
    if summary.sheets_only_in_email:
        print(f"  첨부에만 있는 시트      : {summary.sheets_only_in_email}")
    if summary.sheets_only_in_tracker:
        print(f"  트래커에만 있는 시트    : {summary.sheets_only_in_tracker}")

    if summary.added:
        # 시트별로 묶어서 보여주기
        from collections import Counter
        by_sheet = Counter(d.sheet for d in summary.added)
        print("\n  [추가 후보 - 시트별 개수]")
        for sh, n in sorted(by_sheet.items()):
            print(f"     {sh:10s} : {n}")
        print(f"\n  [추가 후보 - 상위 {max_show}건 미리보기]")
        for d in summary.added[:max_show]:
            print(f"     + {d.sheet} {d.cell} = {d.new!r}")

    if summary.conflicts:
        from collections import Counter
        by_sheet = Counter(d.sheet for d in summary.conflicts)
        print("\n  [충돌 - 시트별 개수]")
        for sh, n in sorted(by_sheet.items()):
            print(f"     {sh:10s} : {n}")
        print(f"\n  [충돌 - 상위 {max_show}건 미리보기]")
        for d in summary.conflicts[:max_show]:
            print(f"     ≠ {d.sheet} {d.cell} : ours={d.old!r}  /  email={d.new!r}")
    print()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("email_xlsx", help="이메일로 받은 M-day Tracking Sheet .xlsx 경로")
    ap.add_argument("--tracker", default=EXCEL_PATH, help="대상 트래커 경로 (기본: config.EXCEL_PATH)")
    ap.add_argument("--apply", action="store_true", help="실제 머지 적용 (기본은 dry-run)")
    ap.add_argument("--max-show", type=int, default=20, help="미리보기 행 수")
    args = ap.parse_args()

    email_path = Path(args.email_xlsx).resolve()
    tracker_path = Path(args.tracker).resolve()
    if not email_path.exists():
        print(f"[FAIL] 첨부 파일 없음: {email_path}", file=sys.stderr)
        return 1
    if not tracker_path.exists():
        print(f"[FAIL] 트래커 없음: {tracker_path}", file=sys.stderr)
        return 1

    print(f"  첨부 : {email_path.name}")
    print(f"  트래커: {tracker_path.name}")

    summary = compare(email_path, tracker_path)
    print_summary(summary, max_show=args.max_show)

    if args.apply:
        if not summary.added:
            print("[INFO] 추가 후보가 0건이라 머지할 게 없습니다.")
            return 0
        backup = apply_merge(email_path, tracker_path, summary)
        print(f"[OK] {len(summary.added)}개 셀 추가됨.")
        print(f"     백업: {backup}")
        if summary.conflicts:
            print(f"\n[!] 충돌 {len(summary.conflicts)}건은 자동 변경하지 않았습니다.")
            print(f"    사람이 직접 확인 후 결정해야 합니다 (--max-show를 늘려서 전체 확인 가능).")
    else:
        print("[INFO] --apply 없이 실행돼 dry-run 모드입니다. 실제 적용하려면 --apply 추가.")

    return 0


if __name__ == "__main__":
    sys.exit(main())

"""
Mail_Log.xlsx 읽기/쓰기 모듈.

핵심 기능:
- 기존 메일 ID(또는 fingerprint) 집합 로딩 — 중복 체크
- 신규 행 append
- Run_Log 시트에 실행 결과 1행 append
- 파일 잠김(다른 프로세스가 엑셀에서 열어둠) 시 백업본에 쓰는 fallback
"""
from __future__ import annotations

import logging
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook

from mail_log_schema import (
    MAIL_LOG_COLUMNS, MAIL_LOG_COL_INDEX, MAIL_LOG_SHEET,
    RUN_LOG_COLUMNS, RUN_LOG_SHEET, MAIL_LOG_XLSX,
    create_empty_template,
)

logger = logging.getLogger("mail_log_excel")


# ============================================================
# 데이터 모델
# ============================================================
@dataclass
class MailRow:
    """Mail_Log 시트에 들어갈 한 행. 모든 필드는 사람 친화 표준 형식."""
    mail_id: str                # 메일ID (Message-ID 또는 UID 기반)
    received_at: str            # 수신일시 "YYYY-MM-DD HH:mm:ss"
    sender_name: str
    sender_email: str
    company: str
    subject: str
    work_type: str
    summary: str
    request: str
    amount: object              # number 또는 ""
    quantity: object            # number 또는 ""
    due_date: str               # "YYYY-MM-DD" 또는 ""
    owner: str
    priority: str
    status: str
    attachments: str            # 콤마 구분
    ref: str                    # IMAP UID 등
    processed_at: str           # "YYYY-MM-DD HH:mm:ss"
    confidence: str
    note: str

    def to_row(self) -> list:
        """MAIL_LOG_COLUMNS 순서대로 정렬된 리스트로 변환."""
        m = {
            "메일ID": self.mail_id,
            "수신일시": self.received_at,
            "발신자이름": self.sender_name,
            "발신자이메일": self.sender_email,
            "회사명": self.company,
            "제목": self.subject,
            "업무유형": self.work_type,
            "핵심내용요약": self.summary,
            "요청사항": self.request,
            "금액": self.amount,
            "수량": self.quantity,
            "납기일": self.due_date,
            "담당자": self.owner,
            "우선순위": self.priority,
            "상태": self.status,
            "첨부파일명": self.attachments,
            "원문참조": self.ref,
            "처리일시": self.processed_at,
            "신뢰도": self.confidence,
            "비고": self.note,
        }
        return [m[name] for name in MAIL_LOG_COLUMNS]


@dataclass
class RunLogRow:
    run_id: str
    started_at: str
    ended_at: str
    checked: int
    new_count: int
    duplicate: int
    review_needed: int
    non_business: int
    has_error: bool
    error_detail: str
    dashboard_status: str

    def to_row(self) -> list:
        m = {
            "실행ID": self.run_id,
            "시작시각": self.started_at,
            "종료시각": self.ended_at,
            "확인메일수": self.checked,
            "신규반영": self.new_count,
            "중복제외": self.duplicate,
            "검토필요": self.review_needed,
            "비업무제외": self.non_business,
            "오류여부": "Y" if self.has_error else "N",
            "오류상세": self.error_detail,
            "대시보드상태": self.dashboard_status,
        }
        return [m[name] for name in RUN_LOG_COLUMNS]


# ============================================================
# 중복 키 (메일ID 우선, 없으면 fingerprint)
# ============================================================
def make_fingerprint(received_at: str, sender_email: str, subject: str) -> str:
    """메일ID가 없을 때 사용할 fallback 키."""
    parts = [
        (received_at or "").strip(),
        (sender_email or "").strip().lower(),
        (subject or "").strip(),
    ]
    return "|".join(parts)


# ============================================================
# 파일 보장
# ============================================================
def ensure_workbook(path: Path | str = MAIL_LOG_XLSX) -> Path:
    """파일이 없으면 빈 템플릿 생성. 있으면 그대로."""
    path = Path(path)
    if not path.exists():
        logger.info("Mail_Log.xlsx 없음 - 빈 템플릿 생성")
        create_empty_template(path, force=False)
    return path


# ============================================================
# 기존 메일ID/fingerprint 로딩
# ============================================================
def load_existing_keys(path: Path | str = MAIL_LOG_XLSX) -> tuple[set[str], set[str]]:
    """(메일ID 집합, fingerprint 집합) 반환."""
    path = Path(path)
    ids: set[str] = set()
    fps: set[str] = set()
    if not path.exists():
        return ids, fps
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("기존 파일 로드 실패: %s (빈 키 집합으로 진행)", exc)
        return ids, fps
    if MAIL_LOG_SHEET not in wb.sheetnames:
        wb.close()
        return ids, fps
    ws = wb[MAIL_LOG_SHEET]
    id_col = MAIL_LOG_COL_INDEX["메일ID"] + 1
    recv_col = MAIL_LOG_COL_INDEX["수신일시"] + 1
    from_col = MAIL_LOG_COL_INDEX["발신자이메일"] + 1
    subj_col = MAIL_LOG_COL_INDEX["제목"] + 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        mid = row[id_col - 1]
        if mid:
            ids.add(str(mid).strip())
        recv = row[recv_col - 1]
        frm = row[from_col - 1]
        subj = row[subj_col - 1]
        fp = make_fingerprint(str(recv or ""), str(frm or ""), str(subj or ""))
        if fp.strip("|"):
            fps.add(fp)
    wb.close()
    return ids, fps


# ============================================================
# 행 append (잠금 대응 포함)
# ============================================================
def append_rows(
    rows: Iterable[MailRow],
    run_log: Optional[RunLogRow] = None,
    path: Path | str = MAIL_LOG_XLSX,
) -> tuple[int, Optional[Path]]:
    """Mail_Log + Run_Log에 행 추가. 잠겨 있으면 백업본 생성 후 그곳에 기록.

    Returns:
        (실제 추가된 Mail_Log 행 수, 잠금 회피 시 사용한 fallback 경로 or None)
    """
    path = Path(path)
    ensure_workbook(path)
    rows_list = list(rows)

    actual_path = path
    fallback: Optional[Path] = None
    try:
        wb = load_workbook(path)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = path.with_name(f"{path.stem}_locked_{ts}{path.suffix}")
        logger.warning("Mail_Log.xlsx 잠김 → 임시 사본에 기록: %s", fallback)
        shutil.copy2(path, fallback)
        wb = load_workbook(fallback)
        actual_path = fallback

    ws = wb[MAIL_LOG_SHEET] if MAIL_LOG_SHEET in wb.sheetnames else wb.active
    appended = 0
    for r in rows_list:
        ws.append(r.to_row())
        appended += 1

    if run_log is not None:
        if RUN_LOG_SHEET not in wb.sheetnames:
            wb.create_sheet(title=RUN_LOG_SHEET)
            ws2 = wb[RUN_LOG_SHEET]
            ws2.append(RUN_LOG_COLUMNS)
        else:
            ws2 = wb[RUN_LOG_SHEET]
        ws2.append(run_log.to_row())

    wb.save(actual_path)
    return appended, fallback


# ============================================================
# Mail_Log 전체 읽기 (대시보드 생성용)
# ============================================================
def read_all_rows(path: Path | str = MAIL_LOG_XLSX) -> list[dict]:
    """Mail_Log 시트 전체 행을 dict 리스트로 반환."""
    path = Path(path)
    if not path.exists():
        return []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        logger.error("Mail_Log 읽기 실패: %s", exc)
        return []
    if MAIL_LOG_SHEET not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[MAIL_LOG_SHEET]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        item = {}
        for i, name in enumerate(MAIL_LOG_COLUMNS):
            item[name] = row[i] if i < len(row) else None
        out.append(item)
    wb.close()
    return out

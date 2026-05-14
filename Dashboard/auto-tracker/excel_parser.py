"""
메일 첨부 작업일보(.xlsx/.xlsm) → DailyReport 변환 모듈.

다우오피스로 들어오는 작업일보 첨부 엑셀의 표준 구조를 가정합니다.

가정하는 엑셀 구조 (시트명: '작업일보'):
    C2  = 공사명          예) "CP CHW1FC Gap Cover 추가 공사"
    C3  = 작업일자         예) "2026.04.28 (화)"
    B7  = 금일 작업내용 텍스트 블록 (라인별/작업종류별 수량)
    E38 = 관리자 (오늘)
    E39 = 작업자 (오늘)
    E40 = 신호수 (오늘)
    E41 = 화기감시자 (오늘)
    E42 = 용접작업자 (오늘)

B7 텍스트 블록 형식 예시:
    - Gap Cover 추가 공사 ( PBS #1 Line ) (진행중)
       -> 37 sets 설치 / 누적 250 sets / 총 395 sets / 63%
       -> TP : #168 ~ #206
       -> NTP : #168 ~ #206
    - Tack 틈새 보완 ( CBS Line )
       ->  0 m 설치 / 누적 0 m / 총 78 m / 0%
"""
from __future__ import annotations

import re
import logging
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from config import SITE_MAPPING
from parser import DailyReport, WorkItem, RawDetail, normalize


logger = logging.getLogger(__name__)


# 작업종류 한글/영문 → 정규화 키 매핑
WORK_TYPE_NORMALIZE = {
    "gapcover": "gapcover",
    "갭커버": "gapcover",
    "cover": "cover",
    "커버": "cover",
    "gapplate": "gapplate",
    "갭플레이트": "gapplate",
    "플레이트": "gapplate",
    "cable": "cable",
    "harness": "cable",
    "harnesscable": "cable",
    "케이블": "cable",
    "하네스케이블": "cable",
    "tack": "tack",
}


def normalize_line(line_text: str) -> str:
    """
    라인 표기를 ROW_MAPPING 키와 매칭 가능한 형태로 정규화합니다.

    예)
        "PBS #1 Line"      → "pb#1line"
        "PBS #01"          → "pb#01line"
        "CBS Line"         → "cbline"
        "ST #2 Line"       → "st#2line"
        "PBS Line 상단부"   → "pb#1line"   (상단=#1)
        "PBS Line 하단부"   → "pb#2line"   (하단=#2)
    """
    s = line_text.strip()
    s = re.sub(r"\s+", "", s).lower()
    # 한글 상단/하단 표기를 #1/#2로 정규화 (현장에서 한 라인을 상·하단으로 부르는 케이스)
    s = s.replace("상단부", "#1").replace("상단", "#1")
    s = s.replace("하단부", "#2").replace("하단", "#2")
    # 'line' 토큰을 모두 제거한 뒤 끝에 한 번만 붙여서 위치 일관성 유지
    # (예: "PBS Line 상단부" → "pbsline#1" → "pbs#1" → "pb#1" → "pb#1line")
    s = s.replace("line", "")
    # PBS → PB
    s = s.replace("pbs", "pb")
    # CBS → CB (line 토큰을 이미 제거했으므로 끝에 cbs가 올 수 있음)
    s = re.sub(r"cbs(?![a-z0-9])", "cb", s)
    s = s + "line"
    return s


def parse_site_code(project_name: str) -> Optional[str]:
    """공사명에서 현장 코드(CHW1FC 등)를 추출합니다."""
    if not project_name:
        return None

    # 1. 직접 사이트 코드 패턴 매칭 (영문3 + 숫자1 + 영문2)
    m = re.search(r"\b([A-Z]{3}\d[A-Z]{2})\b", project_name.upper())
    if m:
        candidate = m.group(1)
        if candidate in SITE_MAPPING.values():
            return candidate

    # 2. 한글 명칭 매칭
    norm = normalize(project_name)
    for key, sheet in SITE_MAPPING.items():
        if key in norm:
            return sheet

    return None


def parse_work_date(date_text) -> Optional[date]:
    """'2026.04.28 (화)' 같은 문자열 또는 date 객체 → date 객체."""
    if date_text is None:
        return None
    # 이미 date/datetime 인스턴스
    if isinstance(date_text, date):
        return date_text if not hasattr(date_text, "date") else date_text.date()

    s = str(date_text).strip()
    # YYYY.MM.DD
    m = re.search(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    # YY.MM.DD
    m = re.search(r"(\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})", s)
    if m:
        try:
            y = int(m.group(1))
            y = 2000 + y if y < 50 else 1900 + y
            return date(y, int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def parse_work_items_from_text(text: str) -> list[WorkItem]:
    """B7 셀의 작업내용 텍스트 블록에서 WorkItem 리스트를 추출합니다."""
    items: list[WorkItem] = []
    if not text:
        return items

    lines = [ln.rstrip() for ln in str(text).splitlines()]

    current_line_name: Optional[str] = None
    current_work_type_raw: Optional[str] = None

    item_header_re = re.compile(
        r"^\s*-\s*(?P<work>[^\(]+?)\s*\(\s*(?P<line>[^)]+?)\s*\)"
    )
    quantity_re = re.compile(
        r"->\s*(?P<qty>[\d,]+(?:\.\d+)?)\s*(?P<unit>[A-Za-z가-힣]+)?\s*설치"
    )
    range_re = re.compile(
        r"->\s*(?P<tag>TP|NTP)\s*:\s*(?P<range>[#\d\s~\-]+)"
    )

    raw_details: list[RawDetail] = []
    last_qty: Optional[float] = None

    def flush_item():
        nonlocal current_line_name, current_work_type_raw, raw_details, last_qty
        if (current_line_name and current_work_type_raw
                and last_qty is not None and last_qty > 0):
            work_norm = normalize(current_work_type_raw)
            work_key = None
            for k, v in WORK_TYPE_NORMALIZE.items():
                if k in work_norm:
                    work_key = v
                    break
            if work_key is None:
                work_key = work_norm[:20]

            line_norm = normalize_line(current_line_name)
            items.append(WorkItem(
                line_raw=current_line_name,
                work_type_raw=current_work_type_raw,
                line_normalized=line_norm,
                work_type=work_key,
                quantity=last_qty,
                detail=", ".join(
                    f"{d.tp_or_ntp} {d.number_range}".strip()
                    for d in raw_details
                ),
                raw_details=raw_details.copy(),
            ))
        raw_details.clear()
        last_qty = None

    for ln in lines:
        m = item_header_re.match(ln)
        if m:
            flush_item()
            current_work_type_raw = m.group("work").strip()
            current_line_name = m.group("line").strip()
            continue

        m = quantity_re.search(ln)
        if m:
            try:
                last_qty = float(m.group("qty").replace(",", ""))
            except ValueError:
                pass
            continue

        m = range_re.search(ln)
        if m:
            raw_details.append(RawDetail(
                tp_or_ntp=m.group("tag"),
                number_range=m.group("range").strip(),
                quantity=0,
            ))
            continue

    flush_item()
    return items


def parse_attachment(file_path) -> Optional[DailyReport]:
    """첨부 엑셀 파일 1개 → DailyReport. 실패 시 None."""
    path = Path(file_path)
    if not path.exists():
        logger.error("파일이 존재하지 않습니다: %s", path)
        return None

    try:
        wb = load_workbook(path, data_only=True, keep_vba=False)
    except Exception as exc:
        logger.error("엑셀 로드 실패: %s — %s", path.name, exc)
        return None

    ws = wb["작업일보"] if "작업일보" in wb.sheetnames else wb[wb.sheetnames[0]]

    project_name = ws["C2"].value or ""
    date_text = ws["C3"].value or ""

    site_code = parse_site_code(str(project_name))
    work_date = parse_work_date(date_text)

    # 파일명에 날짜가 있으면 그걸 우선한다.
    # (작성자가 파일 복사 후 C3을 갱신 안 한 실제 케이스가 있어서 파일명이 더 신뢰 가능)
    fname_date = parse_work_date(path.stem)
    if fname_date:
        if work_date and work_date != fname_date:
            logger.warning(
                "C3 날짜와 파일명 날짜 불일치 → 파일명 사용: C3=%s, 파일명=%s, 파일=%s",
                work_date, fname_date, path.name,
            )
        work_date = fname_date

    if not site_code:
        logger.warning("공사명에서 현장 코드 추출 실패: %r", project_name)
    if not work_date:
        logger.warning("작업일자 파싱 실패: %r", date_text)

    def _to_int(v) -> int:
        try:
            return int(v) if v is not None else 0
        except (ValueError, TypeError):
            return 0

    manager_count = _to_int(ws["E38"].value)
    worker_count = _to_int(ws["E39"].value)
    signal_count = _to_int(ws["E40"].value)
    firewatch_count = _to_int(ws["E41"].value)
    welder_count = _to_int(ws["E42"].value)

    total_workers = (
        manager_count + worker_count + signal_count
        + firewatch_count + welder_count
    )

    work_text = ws["B7"].value or ""
    items = parse_work_items_from_text(str(work_text))

    return DailyReport(
        site_raw=str(project_name),
        site_sheet=site_code or "",
        work_date=work_date,
        total_workers=total_workers if total_workers > 0 else None,
        manager_count=manager_count or None,
        worker_count=worker_count or None,
        items=items,
        raw_message=f"[Excel] {path.name}",
    )
